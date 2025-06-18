from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QLabel, QVBoxLayout, QWidget, QLabel
from PyQt5.QtCore import QTime
import sys
import os
from PyQt5 import uic, QtWidgets
from datetime import datetime
import openpyxl
from logic.logic_handle import AutoCompleterComboBox
from openpyxl.utils import range_boundaries
import copy
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

class StartTaskWindow(QMainWindow):
	def __init__(self, name_of_worker):
		super().__init__()

		uiFilePath = os.path.join(os.path.dirname(__file__), 'startTaskWindow.ui')
		uic.loadUi(uiFilePath, self)
		self.setWindowTitle("Start Task")
		self.worker_name_label = self.findChild(QtWidgets.QLabel, "label")
		self.worker_name_label.setText(name_of_worker)

		self.projects_box = self.findChild(QtWidgets.QComboBox, "comboBox_2")
		self.steps_box = self.findChild(QtWidgets.QComboBox, "comboBox_3")
		self.time_edit = self.findChild(QtWidgets.QTimeEdit, "timeEdit")
		self.time_edit.setTime(QTime.currentTime())
		self.time_edit.setDisplayFormat("HH:mm")           

		self.start_task_button = self.findChild(QtWidgets.QPushButton, "pushButton")
		self.start_task_button.clicked.connect(self.on_start_task)

		self.outputReport_path = os.path.join(os.path.dirname(__file__), '..', 'output_report.xlsx')
		self.inputData_path = os.path.join(os.path.dirname(__file__), '..', 'inputdata.xlsx')

		self.startDataRow = 8 #loai bo cac header va phan trang tri

		projects = self.get_project_from_excel(self.inputData_path)
		AutoCompleterComboBox(self.projects_box, projects)

		self.projects_box.currentTextChanged.connect(self.update_task_box_by_project)
		
	def on_start_task(self):
	   
		name_of_worker = self.worker_name_label.text()
		project = self.projects_box.currentText()
		step = self.steps_box.currentText()
		start_time = self.time_edit.time().toString("HH:mm")

		self.write_start_task_to_excel(self.outputReport_path, name_of_worker, project, step, start_time)

	def write_start_task_to_excel(self, file_path, name_of_worker, project_name, step, start_time):
		wb = openpyxl.load_workbook(file_path)
		#Mo hoac tao mot sheet co ten la ngay bao cao
		today = datetime.today().strftime("%d.%m")
		if today in wb.sheetnames:
			ws = wb[today]
		else:
			ws = self.get_or_create_today_sheet(wb)

		if ws['B8'].value is None:
			row = 8
			ws.cell(row=row, column=2).value= name_of_worker
			ws.cell(row=row,column=3).value = project_name
			ws.cell(row=row,column=4).value = step
			ws.cell(row=row,column=6).value = start_time
			wb.save(file_path)
			self.close()
			return
		else:
			first_row, total_cells = self.find_worker_name_area(name_of_worker, self.outputReport_path)
			if first_row == None: #chua co ten trong report:
				rows_with_data = [row for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=3).value is not None]
				last_row = max(rows_with_data) if rows_with_data else 0
				row = last_row + 2 #tim hang cuoi cung co du lieu o cot C
				ws.cell(row=row, column=2).value= name_of_worker
				ws.cell(row=row,column=3).value = project_name
				ws.cell(row=row,column=4).value = step
				ws.cell(row=row,column=6).value = start_time
				wb.save(file_path)
				self.close()
				return

			row = first_row + total_cells

			#phai xu ly cac khu vuc da mearge
			old_merges = list(ws.merged_cells.ranges)
			ws.merged_cells.ranges = []

			ws.insert_rows(row)

			for cr in old_merges:
			    bounds = CellRange(str(cr))
			    if bounds.min_row >= row:
			        # Nếu merge nằm dưới dòng chèn → dịch xuống
			        bounds.shift(0, 1)  # shift(col_offset, row_offset)
			    ws.merge_cells(str(bounds))

			ws.cell(row=row,column=3).value = project_name
			ws.cell(row=row,column=4).value = step
			ws.cell(row=row,column=6).value = start_time

			#merge cac o lai vs nhau
			start_row = first_row
			end_row = row
			column_letter = 'B'

			unmerge_range = f"{column_letter}{start_row}:{column_letter}{end_row - 1}"
			merge_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

			if total_cells != 1:
				ws.unmerge_cells(str(unmerge_range))
			ws.merge_cells(merge_range)
			wb.save(file_path)
			self.close()

	def find_worker_name_area(self, target_value, file_path):
		wb = openpyxl.load_workbook(file_path)
		today = datetime.today().strftime("%d.%m")

		if today in wb.sheetnames:
			ws = wb[today]
		else:
			ws = self.get_or_create_today_sheet(wb)

		for row in range(1, ws.max_row + 1):
			cell_value = ws[f"B{row}"].value
			if cell_value and str(cell_value).strip() == target_value:
				total_cells = 1  # mặc định ô đơn
				for merged_range in ws.merged_cells.ranges:
					min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
					# Kiểm tra vùng merge bắt đầu đúng vị trí row, cột B(=2)
					if min_row == row and min_col == 2:
						row_count = max_row - min_row + 1
						col_count = max_col - min_col + 1
						total_cells = row_count * col_count
						break  # đã tìm thấy vùng merge phù hợp, thoát vòng for
				return row, total_cells  # trả về dòng đầu tiên và số ô trong vùng (merged hoặc đơn)

		return None, 0  # Không tìm thấy target_value trong cột B


	def get_project_from_excel(self, file_path):
		workbook = openpyxl.load_workbook(file_path)
		sheet = workbook["Project"]
		projects = []

		for row in sheet.iter_rows(min_row=2, max_col=1):
			cell = row[0]
			if cell.value:
				projects.append(str(cell.value))
		return projects

	def get_project_step_from_excel(self, file_path, target_project):
		workbook = openpyxl.load_workbook(file_path)
		sheet = workbook["Project"]
		tasks = []

		for row in sheet.iter_rows(min_row=2):
			if row[0].value == target_project:
				row_data = [cell.value for cell in row[1:]]
				return row_data
  
	def update_task_box_by_project(self, project_name):
		tasks = self.get_project_step_from_excel(self.inputData_path, project_name)

		self.steps_box.clear()
		AutoCompleterComboBox(self.steps_box, tasks)

	def get_or_create_today_sheet(self, wb):
		today = datetime.today().strftime("%d.%m")

		if today in wb.sheetnames:
			return wb[today]

		if "sample" not in wb.sheetnames:
			raise ValueError("Không tìm thấy sheet 'sample' để sao chép.")

		# Dùng hàm built-in để copy
		sample_sheet = wb["sample"]
		new_sheet = wb.copy_worksheet(sample_sheet)
		new_sheet.title = today

		return new_sheet
