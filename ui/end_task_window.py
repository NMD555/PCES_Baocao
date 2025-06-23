from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QLabel, QVBoxLayout, QHBoxLayout, QWidget, QLabel, QLineEdit, QSizePolicy, QMessageBox, QTimeEdit
from PyQt5.QtCore import QTime, QTimer, Qt
import sys
import os
from PyQt5 import uic, QtWidgets
from datetime import datetime
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import openpyxl
from logic.logic_handle import AutoCompleterComboBox
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from datetime import datetime, time as dtime


class EndTaskWindow(QMainWindow):
    def __init__(self, name_of_worker):
        super().__init__()
        self.setWindowTitle("Finish Task")
        self.outputReport_path = os.path.join(os.path.dirname(__file__), '..', 'output_report.xlsx')
        self.task_inputs = []
        self.time_finish = []
        not_finish_task = self.get_unfinish_task(name_of_worker)
        if not not_finish_task:
            QMessageBox.warning(self, "Thông báo", f"{name_of_worker} chưa bắt đầu công việc nào!")
            QTimer.singleShot(0, self.close)  # Đóng sau khi event loop khởi động
            return
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)  # khoảng cách giữa các HBox
        main_layout.setContentsMargins(0, 0, 0, 0)  # (left, top, right, bottom)

        central_widget.setLayout(main_layout)

        name_label = QLabel(name_of_worker)
        name_label.setStyleSheet("font-size: 50px;")
        main_layout.addWidget(name_label)

        for row, col_c, col_d in not_finish_task:
            h_layout = QHBoxLayout()
            h_layout.setSpacing(10)
            h_layout.setContentsMargins(0, 0, 0, 0)


            project_label = QLabel(col_c)
            task_label = QLabel(col_d)
            input_field = QLineEdit()
            finish_time = QTimeEdit()
            finish_time.setTime(QTime.currentTime())
            finish_time.setDisplayFormat("HH:mm")

            self.task_inputs.append((row, input_field))
            self.time_finish.append((row, finish_time))

            project_label.setStyleSheet("font-size: 20px;")
            task_label.setStyleSheet("font-size: 20px;")
            input_field.setStyleSheet("font-size: 20px;")
            finish_time.setStyleSheet("font-size:20px")

            project_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            task_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            input_field.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            finish_time.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            h_layout.addWidget(project_label)
            h_layout.addWidget(task_label)
            h_layout.addWidget(input_field)
            h_layout.addWidget(finish_time)

            h_layout.setStretch(0, 0) 

            main_layout.addLayout(h_layout)

        end_task_button = QPushButton("End Task")
        end_task_button.setStyleSheet("""
        font-size: 30px;
        padding: 20px 40px;
        background-color: #3498db;
        color: white;
        border: none;
        border-radius: 15px;
        """)
        end_task_button.setMinimumWidth(110)
        end_task_button.setMinimumHeight(20)
        end_task_button.clicked.connect(self.end_task_button_clicked)
        main_layout.addWidget(end_task_button, alignment=Qt.AlignCenter)

    def end_task_button_clicked(self):
        wb = openpyxl.load_workbook(self.outputReport_path)
        today = datetime.today().strftime("%d.%m")

        if today in wb.sheetnames:
            ws = wb[today]
        else:
            ws = wb.create_sheet(title=today)

        for i, (row, amount_widget) in enumerate(self.task_inputs):
            # Xử lý amount
            text = amount_widget.text().strip()
            try:
                amount_value = float(text)
            except ValueError:
                amount_value = None

            ws.cell(row=row, column=8, value=amount_value)

            # Nếu amount không trống thì mới ghi giờ
            if text != "":
                _, time_obj = self.time_finish[i]

                # Trường hợp time_obj là QTimeEdit
                if isinstance(time_obj, QTimeEdit):
                    time_str = time_obj.time().toString("HH:mm")

                # Trường hợp là QTime
                elif hasattr(time_obj, 'toString'):
                    time_str = time_obj.toString("HH:mm")

                # Nếu là chuỗi
                else:
                    time_str = str(time_obj)

                ws.cell(row=row, column=7, value=time_str)

        wb.save(self.outputReport_path)
        self.make_gant_chart()
        self.close()

    def get_unfinish_task(self, name):
        wb = openpyxl.load_workbook(self.outputReport_path)
        today = datetime.today().strftime("%d.%m")

        if today in wb.sheetnames:
            ws = wb[today]
        else:
            ws = wb.create_sheet(title=today)

        firstrow, cell = self.find_worker_name_area(name, self.outputReport_path)
        if not firstrow or not cell:
            return []
        not_finish_task_row = []
        for row in range(firstrow, firstrow + cell):
            cell_value = ws.cell(row=row, column=7).value  # Cột G
            if cell_value is None:
                not_finish_task_row.append(row)

        # Trả về danh sách gồm: (row, cột C value, cột D value)
        not_finish_info = []
        for row in not_finish_task_row:
            col_c = ws.cell(row=row, column=3).value  # Cột C
            col_d = ws.cell(row=row, column=4).value  # Cột D
            not_finish_info.append((row, col_c, col_d))

        return not_finish_info
            

    def find_worker_name_area(self, target_value, file_path):
        wb = openpyxl.load_workbook(file_path)
        today = datetime.today().strftime("%d.%m")

        if today in wb.sheetnames:
            ws = wb[today]
        else:
            ws = wb.create_sheet(title=today)

        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"B{row}"].value
            if cell_value and str(cell_value).strip() == target_value:
                total_cells = 1  # mặc định ô đơn
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    # Kiểm tra vùng merge bắt đầu đúng vị trí row, cột B(=2)
                    if min_row == row and min_col == 2:
                        row_count = max_row - min_row + 1
                        col_count = max_col - min_col + 1
                        total_cells = row_count * col_count
                        break  # đã tìm thấy vùng merge phù hợp, thoát vòng for
                return row, total_cells  # trả về dòng đầu tiên và số ô trong vùng (merged hoặc đơn)

        return None, 0  # Không tìm thấy target_value trong cột B

    def make_gant_chart(self):
        wb = openpyxl.load_workbook(self.outputReport_path)
        today = datetime.today().strftime("%d.%m")

        if today in wb.sheetnames:
            ws = wb[today]
        else:
            ws = wb.create_sheet(title=today)

        start_hour = 8
        end_hour = 18
        step_minutes = 30
        start_col = 10

 
        # === Tạo danh sách mốc thời gian ===
        time_range = []
        t = datetime(2023, 1, 1, start_hour, 0)
        end_time = datetime(2023, 1, 1, end_hour, 0)
        while t <= end_time:
            time_range.append(t)
            t += timedelta(minutes=step_minutes)

        # === Ghi hàng tiêu đề thời gian (bắt đầu từ cột K) ===
        for i, t in enumerate(time_range):
            col_letter = get_column_letter(start_col + i)
            ws[f"{col_letter}7"] = t.strftime("%H:%M")

        # === Xử lý từng dòng dữ liệu ===
        row = 8
        max_excel_row = 200
        while row <= max_excel_row:
            start_cell = ws[f"F{row}"]
            end_cell = ws[f"G{row}"]

            if not start_cell.value or not end_cell.value:
                row += 1
                continue

            try:
                start_time = self.parse_time(start_cell.value)
                end_time = self.parse_time(end_cell.value)
            except ValueError:
                row += 1
                continue

            for i, t in enumerate(time_range):
                t_next = t + timedelta(minutes=step_minutes)
                if start_time < t_next.time() and end_time > t.time():
                    col_letter = get_column_letter(start_col + i)
                    cell = ws[f"{col_letter}{row}"]
                    cell.fill = PatternFill(start_color="FF00C0FF", end_color="FF00C0FF", fill_type="solid")


            row += 1

        # === Lưu file mới ===
        wb.save(self.outputReport_path)

    def parse_time(self, cell_val):
        if isinstance(cell_val, dtime):
            return cell_val
        elif isinstance(cell_val, datetime):
            return cell_val.time()
        elif isinstance(cell_val, str):
            try:
                return datetime.strptime(cell_val.strip(), "%H:%M").time()
            except:
                return None
        else:
            return None



