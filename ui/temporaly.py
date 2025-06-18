from PyQt5.QtWidgets import QMainWindow, QWidget, QVBoxLayout, QCompleter
from PyQt5 import QtWidgets

import openpyxl

from PyQt5 import uic
import os

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        uiFilePath = os.path.join(os.path.dirname(__file__), 'ui.ui')
        uic.loadUi(uiFilePath, self)

        self.worker_name_box = self.findChild(QtWidgets.QComboBox, "comboBox")
        self.project_name_box = self.findChild(QtWidgets.QComboBox, "comboBox_2")
        self.task_name_box = self.findChild(QtWidgets.QComboBox, "comboBox_3")

        excel_path = os.path.join(os.path.dirname(__file__),'..', 'inputdata.xlsx')

        names = get_names_from_excel(excel_path)
        projects = get_project_from_excel(excel_path)
        tasks = get_project_step_from_excel(excel_path, "LLS06")

        # Gan completer vao o name
        completer_worker_name = QCompleter(names)
        completer_worker_name.setCaseSensitivity(False)  # không phân biệt hoa thường
        self.worker_name_box.setEditable(True)  # Phải cho phép người dùng gõ
        self.worker_name_box.setCompleter(completer_worker_name)
        # Optionally: add sẵn các tên vào dropdown
        self.worker_name_box.addItems(names)

        #Gan completer vao o project
        completer_project = QCompleter(projects)
        completer_project.setCaseSensitivity(False)  # không phân biệt hoa thường
        self.project_name_box.setEditable(True)  # Phải cho phép người dùng gõ
        self.project_name_box.setCompleter(completer_project)
        # Optionally: add sẵn các tên vào dropdown
        self.project_name_box.addItems(projects)

        self.project_name_box.currentTextchanged.connect(self.update_task_box_by_project)

def get_names_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Name"]
    names = []

    for row in sheet.iter_rows(min_row=2, max_col=1):  # Bỏ qua tiêu đề
        cell = row[0]
        if cell.value:
            names.append(str(cell.value))
    return names

def get_project_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Project"]
    projects = []

    for row in sheet.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value:
            projects.append(str(cell.value))
    return projects

def get_project_step_from_excel(file_path, target_project):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Project"]
    tasks = []

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == target_project:
            row_data = [cell.value for cell in row]
            return row_data
  
def update_task_box_by_project(self, project_name):
    tasks = get_project_step_from_excel(excel_path, project_name)

    self.task_name_box.clear()
    completer_task = QCompleter(tasks)
    completer_task.setCaseSensitivity(False)

        